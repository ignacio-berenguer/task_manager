"""Excel export for the documentos_items table."""

import sqlite3
import logging
from copy import copy
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.config import settings as config

logger = logging.getLogger('portfolio_export')

# Database column -> Excel column mapping
DOCUMENTOS_ITEMS_COLUMN_MAPPING = {
    'id': 'ID',
    'portfolio_id': 'Portfolio ID',
    'nombre_fichero': 'Nombre Fichero',
    'tipo_documento': 'Tipo Documento',
    'tipo_registro': 'Tipo Registro',
    'texto': 'Texto',
}


def export_documentos_items(db_path: str = None) -> dict:
    """
    Export documentos_items table to Excel.

    Args:
        db_path: Path to SQLite database (uses config default if not specified)

    Returns:
        dict with export statistics
    """
    db_path = db_path or config.DATABASE_PATH
    output_path = Path(config.DOCUMENTOS_ITEMS_EXPORT_PATH)

    logger.info(f"Starting documentos_items export to {output_path}")
    logger.info(f"Worksheet: {config.DOCUMENTOS_ITEMS_EXPORT_WORKSHEET}")
    logger.info(f"Table: {config.DOCUMENTOS_ITEMS_EXPORT_TABLE}")

    # Read data from database
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM documentos_items ORDER BY portfolio_id, nombre_fichero, tipo_registro")
    rows = cursor.fetchall()
    db_columns = [description[0] for description in cursor.description]
    conn.close()

    logger.info(f"Read {len(rows)} rows from documentos_items table")

    # Check if we can use existing template or need to create from scratch
    use_template = False
    if output_path.exists():
        is_xlsm = output_path.suffix.lower() == '.xlsm'
        try:
            wb = load_workbook(output_path, keep_vba=is_xlsm)
            if config.DOCUMENTOS_ITEMS_EXPORT_WORKSHEET in wb.sheetnames:
                ws = wb[config.DOCUMENTOS_ITEMS_EXPORT_WORKSHEET]
                for t in ws.tables.values():
                    if t.name == config.DOCUMENTOS_ITEMS_EXPORT_TABLE:
                        use_template = True
                        break
            if not use_template:
                wb.close()
        except PermissionError:
            error_msg = f"Cannot open Excel file: {output_path}. The file may be open in Excel. Please close it and try again."
            logger.error(error_msg)
            raise PermissionError(error_msg)

    if use_template:
        return _export_to_template(wb, ws, rows, db_columns, output_path)
    else:
        return _export_new_file(rows, db_columns, output_path)


def _export_new_file(rows, db_columns, output_path):
    """Create a new Excel file with documentos_items data."""
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    logger.info(f"Creating new Excel file: {output_path}")

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = config.DOCUMENTOS_ITEMS_EXPORT_WORKSHEET

    # Write headers
    export_cols = list(DOCUMENTOS_ITEMS_COLUMN_MAPPING.keys())
    headers = [DOCUMENTOS_ITEMS_COLUMN_MAPPING[c] for c in export_cols]
    ws.append(headers)

    # Write data
    for db_row in rows:
        row_dict = dict(db_row)
        ws.append([row_dict.get(c) for c in export_cols])

    # Create Excel Table
    max_row = len(rows) + 1  # header + data
    max_col = len(headers)
    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=config.DOCUMENTOS_ITEMS_EXPORT_TABLE, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 12)

    try:
        wb.save(output_path)
        logger.info(f"Saved new workbook to {output_path}")
    except PermissionError:
        error_msg = f"Cannot save Excel file: {output_path}. Please close it and try again."
        logger.error(error_msg)
        raise PermissionError(error_msg)

    result = {
        'rows_exported': len(rows),
        'columns_exported': len(headers),
        'output_file': str(output_path),
        'timestamp': datetime.now().isoformat(),
    }
    logger.info(f"Documentos items export complete: {result}")
    return result


def _export_to_template(wb, ws, rows, db_columns, output_path):
    """Export documentos_items into an existing Excel template with a table."""
    from openpyxl.utils.cell import range_boundaries

    table = None
    for t in ws.tables.values():
        if t.name == config.DOCUMENTOS_ITEMS_EXPORT_TABLE:
            table = t
            break

    logger.info(f"Found table '{table.name}' with ref: {table.ref}")

    # Parse table reference
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    # Get header row
    header_row = min_row
    excel_headers = []
    for col in range(min_col, max_col + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        excel_headers.append(cell_value)

    logger.info(f"Excel table has {len(excel_headers)} columns")

    # Build reverse mapping
    excel_to_db_mapping = {v.lower(): k for k, v in DOCUMENTOS_ITEMS_COLUMN_MAPPING.items()}

    # Find exportable columns
    export_columns = []
    for col_idx, excel_header in enumerate(excel_headers):
        if excel_header is None:
            continue
        header_lower = excel_header.lower()
        if header_lower in excel_to_db_mapping:
            db_col = excel_to_db_mapping[header_lower]
            if db_col in db_columns:
                export_columns.append((col_idx, db_col))

    logger.info(f"Will export {len(export_columns)} columns")

    # Capture cell formats from first data row
    data_start_row = header_row + 1
    old_data_rows = max_row - header_row

    column_formats = {}
    if old_data_rows > 0:
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=data_start_row, column=col)
            column_formats[col] = {
                'number_format': cell.number_format,
                'font': copy(cell.font) if cell.font else None,
                'alignment': copy(cell.alignment) if cell.alignment else None,
            }

    # Clear existing data rows
    logger.info(f"Clearing {old_data_rows} existing data rows")
    for row in range(data_start_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = None

    # Write new data rows
    logger.info(f"Writing {len(rows)} new data rows")
    for row_idx, db_row in enumerate(rows):
        excel_row = data_start_row + row_idx
        row_dict = dict(db_row)

        for excel_col_offset, db_col_name in export_columns:
            excel_col = min_col + excel_col_offset
            value = row_dict.get(db_col_name)

            cell = ws.cell(row=excel_row, column=excel_col)
            cell.value = value

            if excel_col in column_formats:
                fmt = column_formats[excel_col]
                if fmt['number_format']:
                    cell.number_format = fmt['number_format']
                if fmt['font']:
                    cell.font = copy(fmt['font'])
                if fmt['alignment']:
                    cell.alignment = copy(fmt['alignment'])

    # Resize table
    new_max_row = data_start_row + len(rows) - 1 if rows else header_row
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
    logger.info(f"Resizing table from {table.ref} to {new_ref}")
    table.ref = new_ref

    # Save workbook
    try:
        wb.save(output_path)
        logger.info(f"Saved workbook to {output_path}")
    except PermissionError:
        error_msg = f"Cannot save Excel file: {output_path}. Please close it and try again."
        logger.error(error_msg)
        raise PermissionError(error_msg)

    result = {
        'rows_exported': len(rows),
        'columns_exported': len(export_columns),
        'output_file': str(output_path),
        'timestamp': datetime.now().isoformat(),
    }

    logger.info(f"Documentos items export complete: {result}")
    return result
