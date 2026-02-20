"""
Excel export module for datos_relevantes.
Exports database table to Excel table while preserving formatting.
"""

import sqlite3
import logging
from copy import copy
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter

from src.config import settings as config

logger = logging.getLogger('portfolio_export')

# Database column -> Excel column mapping
DATOS_RELEVANTES_COLUMN_MAPPING = {
    'portfolio_id': 'Portfolio ID',
    'nombre': 'Nombre',
    'unidad': 'Unidad',
    'origen': 'Origen',
    'digital_framework_level_1': 'Digital Framework Level 1',
    'prioridad_descriptiva': 'Prioridad Descriptiva',
    'cluster': 'Cluster 2025',
    'priorizacion': 'Priorización',
    'tipo': 'Tipo',
    'referente_negocio': 'Referente Negocio',
    'referente_bi': 'Referente BI',
    'jira_id': 'JIRA ID',
    'it_partner': 'IT Partner',
    'referente_ict': 'Referente ICT',
    'tipo_agrupacion': 'Tipo Agrupación',
    'capex_opex': 'CAPEX-OPEX',
    'cini': 'CINI',
    'fecha_prevista_pes': 'Fecha Prevista PES',
    'estado_de_la_iniciativa': 'Estado de la Iniciativa',
    'fecha_de_ultimo_estado': 'Fecha de último Estado',
    'estado_de_la_iniciativa_2026': 'Estado de la Iniciativa 2026',
    'estado_aprobacion': 'Estado Aprobación',
    'estado_ejecucion': 'Estado Ejecución',
    'estado_agrupado': 'Estado Agrupado',
    'estado_dashboard': 'Estado Dashboard',
    'estado_requisito_legal': 'Estado Requisito Legal',
    'estado_sm100': 'Estado SM100',
    'estado_sm200': 'Estado SM200',
    'iniciativa_aprobada': 'Iniciativa Aprobada',
    'budget_2024': 'Budget 2024',
    'importe_sm200_24': 'Importe SM200 24',
    'importe_aprobado_2024': 'Importe Aprobado 2024',
    'importe_citetic_24': 'Importe CITETIC 24',
    'importe_facturacion_2024': 'Importe Facturación 2024',
    'importe_2024': 'Importe 2024',
    'budget_2025': 'Budget 2025',
    'importe_sm200_2025': 'Importe SM200 2025',
    'importe_aprobado_2025': 'Importe Aprobado 2025',
    'importe_facturacion_2025': 'Importe Facturación 2025',
    'importe_2025': 'Importe 2025',
    'importe_2025_cc_re': 'Importe 2025 CC RE',
    'nuevo_importe_2025': 'Nuevo Importe 2025',
    'budget_2026': 'Budget 2026',
    'importe_sm200_2026': 'Importe SM200 2026',
    'importe_aprobado_2026': 'Importe Aprobado 2026',
    'importe_facturacion_2026': 'Importe Facturación 2026',
    'importe_2026': 'Importe 2026',
    'budget_2027': 'Budget 2027',
    'importe_sm200_2027': 'Importe SM200 2027',
    'importe_aprobado_2027': 'Importe Aprobado 2027',
    'importe_facturacion_2027': 'Importe Facturación 2027',
    'importe_2027': 'Importe 2027',
    'importe_2028': 'Importe 2028',
    'en_presupuesto_del_ano': 'En Presupuesto del Año',
    'calidad_valoracion': 'Calidad Valoración',
    'siguiente_accion': 'Siguiente Acción',
    'esta_en_los_206_me_de_2026': 'Está en los 20,6 M€ de 2026',
    'iniciativa_cerrada_economicamente': 'Iniciativa cerrada económicamente',
    'activo_ejercicio_actual': 'Activo en el ejercicio actual',
    'fecha_sm100': 'Fecha SM100',
    'fecha_aprobada_con_cct': 'Fecha Aprobada con CCT',
    'fecha_en_ejecucion': 'Fecha en Ejecución',
    'fecha_limite': 'Fecha Límite',
    'fecha_limite_comentarios': 'Fecha Límite Comentarios',
    'diferencia_apr_eje_exc_ept': 'Diferencia APR-EJE exc EPT',
    'cluster_de_antes_de_1906': 'Cluster de Antes de 19.06',
}

# Database columns that contain dates stored as text (YYYY-MM-DD format)
DATE_COLUMNS = {
    'fecha_prevista_pes',
    'fecha_de_ultimo_estado',
    'fecha_sm100',
    'fecha_aprobada_con_cct',
    'fecha_en_ejecucion',
    'fecha_limite',
}

# Excel date format
EXCEL_DATE_FORMAT = 'DD/MM/YYYY'


def convert_text_to_date(value: str) -> Optional[datetime]:
    """Convert YYYY-MM-DD text to datetime object for Excel."""
    if not value or not isinstance(value, str):
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d')
    except ValueError:
        return None


def export_datos_relevantes(db_path: str = None) -> dict:  # type: ignore
    """
    Export datos_relevantes table to Excel.

    Args:
        db_path: Path to SQLite database (uses config default if not specified)

    Returns:
        dict with export statistics
    """
    db_path = db_path or config.DATABASE_PATH

    # Build output path
    output_path = Path(config.DATOS_RELEVANTES_PATH)

    logger.info(f"Starting export to {output_path}")
    logger.info(f"Worksheet: {config.DATOS_RELEVANTES_WORKSHEET}")
    logger.info(f"Table: {config.DATOS_RELEVANTES_TABLE}")

    # Check if output file exists
    if not output_path.exists():
        error_msg = f"Output file not found: {output_path}"
        logger.error(error_msg)
        raise FileNotFoundError(error_msg)

    # Read data from database
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM datos_relevantes ORDER BY portfolio_id")
    rows = cursor.fetchall()

    # Get column names from database
    db_columns = [description[0] for description in cursor.description]
    conn.close()

    logger.info(f"Read {len(rows)} rows from database")

    # Load workbook (keep_vba=True only for .xlsm files)
    is_xlsm = output_path.suffix.lower() == '.xlsm'
    try:
        wb = load_workbook(output_path, keep_vba=is_xlsm)
    except PermissionError:
        error_msg = f"Cannot open Excel file: {output_path}. The file may be open in Excel. Please close it and try again."
        logger.error(error_msg)
        raise PermissionError(error_msg)
    except Exception as e:
        error_msg = f"Cannot open Excel file: {output_path}. Error: {e}"
        logger.error(error_msg)
        raise

    # Get worksheet
    if config.DATOS_RELEVANTES_WORKSHEET not in wb.sheetnames:
        error_msg = f"Worksheet '{config.DATOS_RELEVANTES_WORKSHEET}' not found in workbook"
        logger.error(error_msg)
        raise ValueError(error_msg)

    ws = wb[config.DATOS_RELEVANTES_WORKSHEET]

    # Find the table
    table = None
    for t in ws.tables.values():
        if t.name == config.DATOS_RELEVANTES_TABLE:
            table = t
            break

    if not table:
        error_msg = f"Table '{config.DATOS_RELEVANTES_TABLE}' not found in worksheet"
        logger.error(error_msg)
        raise ValueError(error_msg)

    logger.info(f"Found table '{table.name}' with ref: {table.ref}")

    # Parse table reference (e.g., "A1:BZ100")
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    # Get header row (first row of table)
    header_row = min_row
    excel_headers = []
    for col in range(min_col, max_col + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        excel_headers.append(cell_value)

    logger.info(f"Excel table has {len(excel_headers)} columns")

    # Build reverse mapping (Excel column name -> DB column name) - case insensitive
    excel_to_db_mapping = {v.lower(): k for k, v in DATOS_RELEVANTES_COLUMN_MAPPING.items()}

    # Find which columns we can export (columns that exist in both mapping and Excel)
    export_columns = []  # List of (excel_col_idx, db_col_name)
    unmatched_headers = []
    for col_idx, excel_header in enumerate(excel_headers):
        if excel_header is None:
            continue
        header_lower = excel_header.lower()
        if header_lower in excel_to_db_mapping:
            db_col = excel_to_db_mapping[header_lower]
            if db_col in db_columns:
                export_columns.append((col_idx, db_col))
            else:
                logger.warning(f"Column '{excel_header}' maps to '{db_col}' but not found in database")
        else:
            unmatched_headers.append(excel_header)

    if unmatched_headers:
        logger.info(f"Excel columns not in mapping: {unmatched_headers[:10]}{'...' if len(unmatched_headers) > 10 else ''}")

    logger.info(f"Will export {len(export_columns)} columns")

    # Clear existing data rows (keep header row)
    data_start_row = header_row + 1
    old_data_rows = max_row - header_row

    # Capture cell formats from first data row before clearing (number_format, font, alignment only)
    column_formats = {}
    if old_data_rows > 0:
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=data_start_row, column=col)
            column_formats[col] = {
                'number_format': cell.number_format,
                'font': copy(cell.font) if cell.font else None,
                'alignment': copy(cell.alignment) if cell.alignment else None,
            }
        logger.info(f"Captured formats from {len(column_formats)} columns")

    logger.info(f"Clearing {old_data_rows} existing data rows")

    for row in range(data_start_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = None

    # Write new data rows with formatting
    logger.info(f"Writing {len(rows)} new data rows")

    for row_idx, db_row in enumerate(rows):
        excel_row = data_start_row + row_idx
        row_dict = dict(db_row)

        for excel_col_offset, db_col_name in export_columns:
            excel_col = min_col + excel_col_offset
            value = row_dict.get(db_col_name)
            cell = ws.cell(row=excel_row, column=excel_col)

            # Convert date columns from text (YYYY-MM-DD) to Excel date
            if db_col_name in DATE_COLUMNS and value:
                date_value = convert_text_to_date(value)
                if date_value:
                    cell.value = date_value
                    cell.number_format = EXCEL_DATE_FORMAT
                else:
                    cell.value = value  # Keep original if conversion fails
            else:
                cell.value = value

            # Apply format from first row if available (excluding border and fill)
            # Skip number_format for date columns (already set above)
            if excel_col in column_formats:
                fmt = column_formats[excel_col]
                if fmt['number_format'] and db_col_name not in DATE_COLUMNS:
                    cell.number_format = fmt['number_format']
                if fmt['font']:
                    cell.font = copy(fmt['font'])
                if fmt['alignment']:
                    cell.alignment = copy(fmt['alignment'])

    # Resize table to fit new data
    new_max_row = data_start_row + len(rows) - 1 if rows else header_row
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"

    logger.info(f"Resizing table from {table.ref} to {new_ref}")
    table.ref = new_ref

    # Save workbook
    try:
        wb.save(output_path)
        logger.info(f"Saved workbook to {output_path}")
    except PermissionError:
        error_msg = f"Cannot save Excel file: {output_path}. The file may be open in Excel. Please close it and try again."
        logger.error(error_msg)
        raise PermissionError(error_msg)
    except Exception as e:
        error_msg = f"Cannot save Excel file: {output_path}. Error: {e}"
        logger.error(error_msg)
        raise

    result = {
        'rows_exported': len(rows),
        'columns_exported': len(export_columns),
        'output_file': str(output_path),
        'timestamp': datetime.now().isoformat()
    }

    logger.info(f"Export complete: {result}")

    return result


if __name__ == "__main__":
    import sys

    # Setup basic logging for standalone execution
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    db_path = sys.argv[1] if len(sys.argv) > 1 else config.DATABASE_PATH

    try:
        result = export_datos_relevantes(db_path)
        logger.info("Export complete:")
        logger.info(f"  Rows exported: {result['rows_exported']}")
        logger.info(f"  Columns exported: {result['columns_exported']}")
        logger.info(f"  Output file: {result['output_file']}")
    except Exception as e:
        logger.error(f"Error: {e}")
        sys.exit(1)
