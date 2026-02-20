"""Excel export for the documentos table."""

import sqlite3
import logging
from copy import copy
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.config import settings as config

logger = logging.getLogger('portfolio_summarize')

# Date columns that should be converted from ISO string to Excel date
_DATE_COLUMNS = {'fecha_creacion', 'fecha_actualizacion'}

# Database column -> Excel column mapping
DOCUMENTOS_COLUMN_MAPPING = {
    'nombre_fichero': 'Nombre Fichero',
    'portfolio_id': 'Portfolio ID',
    'tipo_documento': 'Tipo Documento',
    'enlace_documento': 'Enlace Documento',
    'estado_proceso_documento': 'Estado Proceso',
    'resumen_documento': 'Resumen Documento',
    'ruta_documento': 'Ruta Documento',
    'fecha_creacion': 'Fecha Creación',
    'fecha_actualizacion': 'Fecha Actualización',
}


def export_documentos(db_path: str = None) -> dict:
    """
    Export documentos table to Excel.

    Args:
        db_path: Path to SQLite database (uses config default if not specified)

    Returns:
        dict with export statistics
    """
    db_path = db_path or config.DATABASE_PATH
    output_path = Path(config.DOCUMENTOS_EXPORT_PATH)

    logger.info(f"Starting documentos export to {output_path}")
    logger.info(f"Worksheet: {config.DOCUMENTOS_EXPORT_WORKSHEET}")
    logger.info(f"Table: {config.DOCUMENTOS_EXPORT_TABLE}")

    if not output_path.exists():
        error_msg = f"Output file not found: {output_path}"
        logger.error(error_msg)
        raise FileNotFoundError(error_msg)

    # Read data from database
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM documentos ORDER BY portfolio_id, tipo_documento")
    rows = cursor.fetchall()
    db_columns = [description[0] for description in cursor.description]
    conn.close()

    logger.info(f"Read {len(rows)} rows from documentos table")

    # Load workbook
    is_xlsm = output_path.suffix.lower() == '.xlsm'
    try:
        wb = load_workbook(output_path, keep_vba=is_xlsm)
    except PermissionError:
        error_msg = f"Cannot open Excel file: {output_path}. The file may be open in Excel. Please close it and try again."
        logger.error(error_msg)
        raise PermissionError(error_msg)

    # Get worksheet
    if config.DOCUMENTOS_EXPORT_WORKSHEET not in wb.sheetnames:
        error_msg = f"Worksheet '{config.DOCUMENTOS_EXPORT_WORKSHEET}' not found in workbook"
        logger.error(error_msg)
        raise ValueError(error_msg)

    ws = wb[config.DOCUMENTOS_EXPORT_WORKSHEET]

    # Find the table
    table = None
    for t in ws.tables.values():
        if t.name == config.DOCUMENTOS_EXPORT_TABLE:
            table = t
            break

    if not table:
        error_msg = f"Table '{config.DOCUMENTOS_EXPORT_TABLE}' not found in worksheet"
        logger.error(error_msg)
        raise ValueError(error_msg)

    logger.info(f"Found table '{table.name}' with ref: {table.ref}")

    # Parse table reference
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    # Get header row
    header_row = min_row
    excel_headers = []
    for col in range(min_col, max_col + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        excel_headers.append(cell_value)

    logger.info(f"Excel table has {len(excel_headers)} columns")

    # Build reverse mapping (Excel column name -> DB column name) - case insensitive
    excel_to_db_mapping = {v.lower(): k for k, v in DOCUMENTOS_COLUMN_MAPPING.items()}

    # Find exportable columns
    export_columns = []
    for col_idx, excel_header in enumerate(excel_headers):
        if excel_header is None:
            continue
        header_lower = excel_header.lower()
        if header_lower in excel_to_db_mapping:
            db_col = excel_to_db_mapping[header_lower]
            if db_col in db_columns:
                export_columns.append((col_idx, db_col))

    logger.info(f"Will export {len(export_columns)} columns")

    # Capture cell formats from first data row
    data_start_row = header_row + 1
    old_data_rows = max_row - header_row

    column_formats = {}
    if old_data_rows > 0:
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=data_start_row, column=col)
            column_formats[col] = {
                'number_format': cell.number_format,
                'font': copy(cell.font) if cell.font else None,
                'alignment': copy(cell.alignment) if cell.alignment else None,
            }

    # Clear existing data rows
    logger.info(f"Clearing {old_data_rows} existing data rows")
    for row in range(data_start_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = None

    # Write new data rows
    logger.info(f"Writing {len(rows)} new data rows")
    for row_idx, db_row in enumerate(rows):
        excel_row = data_start_row + row_idx
        row_dict = dict(db_row)

        for excel_col_offset, db_col_name in export_columns:
            excel_col = min_col + excel_col_offset
            value = row_dict.get(db_col_name)

            # Convert ISO date strings to datetime objects for Excel
            if db_col_name in _DATE_COLUMNS and value and isinstance(value, str):
                try:
                    value = datetime.fromisoformat(value)
                except (ValueError, TypeError):
                    pass  # Keep original string if parsing fails

            cell = ws.cell(row=excel_row, column=excel_col)
            cell.value = value

            # Apply date format for date columns
            if db_col_name in _DATE_COLUMNS and isinstance(value, datetime):
                cell.number_format = 'DD/MM/YYYY'
            elif excel_col in column_formats:
                fmt = column_formats[excel_col]
                if fmt['number_format']:
                    cell.number_format = fmt['number_format']
                if fmt['font']:
                    cell.font = copy(fmt['font'])
                if fmt['alignment']:
                    cell.alignment = copy(fmt['alignment'])

    # Resize table
    new_max_row = data_start_row + len(rows) - 1 if rows else header_row
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
    logger.info(f"Resizing table from {table.ref} to {new_ref}")
    table.ref = new_ref

    # Save workbook
    try:
        wb.save(output_path)
        logger.info(f"Saved workbook to {output_path}")
    except PermissionError:
        error_msg = f"Cannot save Excel file: {output_path}. The file may be open in Excel. Please close it and try again."
        logger.error(error_msg)
        raise PermissionError(error_msg)

    result = {
        'rows_exported': len(rows),
        'columns_exported': len(export_columns),
        'output_file': str(output_path),
        'timestamp': datetime.now().isoformat(),
    }

    logger.info(f"Documentos export complete: {result}")
    return result
